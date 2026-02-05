import json
import base64
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# Style constants
HEADER_FILL = PatternFill(fill_type='solid', fgColor='1e40af')
HEADER_FONT = Font(color='FFFFFF', bold=True)
CHARGE_FILL = PatternFill(fill_type='solid', fgColor='C6EFCE')  # Green(Charge)
DENY_FILL = PatternFill(fill_type='solid', fgColor='FFC7CE')  # Red(Deny)
REVIEW_FILL = PatternFill(fill_type='solid', fgColor='FFEB9C')  # Yellow(Review)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Code categories
INITIAL_CODES = ['99221', '99222', '99223']
SUBSEQUENT_CODES = ['99231', '99232', '99233']
DISCHARGE_CODES = ['99238', '99239']
CRITICAL_CODES = ['99291']
ACP_CODES = ['99497']
UNLISTED_CODES = ['99499']


def apply_charge_scrub(df):
    """
    Apply billing rules to determine CHARGE/DENY/REVIEW for each row.
    Returns tuple of (DataFrame with _Recommendation and _Notes columns, duplicate_count).
    """
    # Initialize columns
    df = df.copy()
    df['_Recommendation'] = 'CHARGE'
    df['_Notes'] = ''

    # Check required columns
    required_cols = ['PATIENTNAME', 'PROCEDURECODE',
                     'SERVICEDATE', 'BILLINGPROVIDER']
    if not all(col in df.columns for col in required_cols):
        return df, 0  # Can't scrub without required columns

    # Rule 0: Detect and mark duplicates (same patient + CPT + DOS)
    duplicate_mask = df.duplicated(
        subset=['PATIENTNAME', 'PROCEDURECODE', 'SERVICEDATE'], keep='first')
    duplicate_count = duplicate_mask.sum()
    df.loc[duplicate_mask, '_Recommendation'] = 'DENY'
    df.loc[duplicate_mask, '_Notes'] = 'Duplicate: same patient, CPT, and DOS'

    # Rule 1: Unlisted codes always DENY
    mask_99499 = df['PROCEDURECODE'].isin(UNLISTED_CODES)
    df.loc[mask_99499, '_Recommendation'] = 'DENY'
    df.loc[mask_99499, '_Notes'] = 'Unlisted E/M code'

    # Process rules 2-8 by patient + DOS groups
    for (patient, dos), group in df.groupby(['PATIENTNAME', 'SERVICEDATE']):
        if len(group) == 1:
            continue  # Single charge, no conflict rules apply

        indices = group.index
        codes = group['PROCEDURECODE'].tolist()

        has_initial = any(c in INITIAL_CODES for c in codes)
        has_subsequent = any(c in SUBSEQUENT_CODES for c in codes)
        has_discharge = any(c in DISCHARGE_CODES for c in codes)
        has_critical = any(c in CRITICAL_CODES for c in codes)
        has_acp = any(c in ACP_CODES for c in codes)

        # Get indices by code type
        initial_idx = [i for i in indices if df.loc[i,
                                                    'PROCEDURECODE'] in INITIAL_CODES]
        subsequent_idx = [i for i in indices if df.loc[i,
                                                       'PROCEDURECODE'] in SUBSEQUENT_CODES]
        discharge_idx = [i for i in indices if df.loc[i,
                                                      'PROCEDURECODE'] in DISCHARGE_CODES]
        critical_idx = [i for i in indices if df.loc[i,
                                                     'PROCEDURECODE'] in CRITICAL_CODES]
        acp_idx = [i for i in indices if df.loc[i,
                                                'PROCEDURECODE'] in ACP_CODES]

        # Rule 2: Initial + Subsequent → CHARGE initial, DENY subsequent
        if has_initial and has_subsequent and not has_critical:
            for i in subsequent_idx:
                if df.loc[i, '_Recommendation'] != 'DENY':
                    df.loc[i, '_Recommendation'] = 'DENY'
                    df.loc[i, '_Notes'] = 'Subsequent with Initial on same DOS (no Critical)'

        # Rule 3: Initial + Discharge → CHARGE initial, DENY discharge
        if has_initial and has_discharge and not has_critical:
            for i in discharge_idx:
                if df.loc[i, '_Recommendation'] != 'DENY':
                    df.loc[i, '_Recommendation'] = 'DENY'
                    df.loc[i, '_Notes'] = 'Discharge with Initial on same DOS (no Critical)'

        # Rule 4: Discharge + Subsequent (no initial) → CHARGE discharge, DENY subsequent
        if has_discharge and has_subsequent and not has_initial and not has_critical:
            for i in subsequent_idx:
                if df.loc[i, '_Recommendation'] != 'DENY':
                    df.loc[i, '_Recommendation'] = 'DENY'
                    df.loc[i,
                           '_Notes'] = 'Subsequent with Discharge (no Initial)'

        # Rule 5: Critical + Subsequent (no initial/discharge) → REVIEW both (timeline dependent)
        if has_critical and has_subsequent and not has_initial and not has_discharge:
            for i in critical_idx + subsequent_idx:
                if df.loc[i, '_Recommendation'] not in ['DENY']:
                    df.loc[i, '_Recommendation'] = 'REVIEW'
                    df.loc[i,
                           '_Notes'] = 'Critical + Subsequent (no Initial/Discharge) (timeline dependent)'

        # Rule 6: Initial + Critical → REVIEW both (timeline dependent)
        if has_initial and has_critical:
            for i in initial_idx + critical_idx:
                if df.loc[i, '_Recommendation'] not in ['DENY']:
                    df.loc[i, '_Recommendation'] = 'REVIEW'
                    df.loc[i,
                           '_Notes'] = 'Initial + Critical (timeline dependent)'

        # Rule 7: Critical + Subsequent + Discharge (no initial) → CHARGE critical, DENY others
        if has_critical and has_subsequent and has_discharge and not has_initial:
            for i in subsequent_idx + discharge_idx:
                if df.loc[i, '_Recommendation'] != 'DENY':
                    df.loc[i, '_Recommendation'] = 'DENY'
                    df.loc[i,
                           '_Notes'] = 'Critical + Subsequent + Discharge (no Initial)'

        # Rule 8: ACP by same billing provider as initial → DENY ACP
        if has_acp and has_initial:
            initial_providers = set(
                df.loc[initial_idx, 'BILLINGPROVIDER'].tolist())
            for i in acp_idx:
                acp_provider = df.loc[i, 'BILLINGPROVIDER']
                if acp_provider in initial_providers:
                    if df.loc[i, '_Recommendation'] != 'DENY':
                        df.loc[i, '_Recommendation'] = 'DENY'
                        df.loc[i, '_Notes'] = 'ACP by same provider as Initial'

    return df, duplicate_count


def create_excel_with_scrub(df, duplicate_count):
    """
    Create Excel workbook with scrub recommendations and duplicate highlighting.
    """
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Scrub Results"

    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_data.cell(row=r_idx, column=c_idx, value=value)
            cell.border = THIN_BORDER

            if r_idx == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT

    # Apply color to key clinical columns based on recommendation
    color_columns = ['PATIENTNAME', 'PROCEDURECODE', 'SERVICEDATE',
                     'BILLINGPROVIDER', '_Recommendation', '_Notes']
    color_col_indices = [df.columns.get_loc(col) + 1 for col in color_columns
                         if col in df.columns]

    for r_idx in range(2, len(df) + 2):  # Start at row 2 (after header)
        data_idx = r_idx - 2
        recommendation = df.iloc[data_idx]['_Recommendation']

        if recommendation == 'DENY':
            fill = DENY_FILL
        elif recommendation == 'REVIEW':
            fill = REVIEW_FILL
        elif recommendation == 'CHARGE':
            fill = CHARGE_FILL
        else:
            continue

        for col_idx in color_col_indices:
            ws_data.cell(row=r_idx, column=col_idx).fill = fill

    # Autofit columns
    for col_idx, column in enumerate(df.columns, 1):
        max_length = len(str(column))
        for cell_value in df[column].astype(str):
            max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 50)
        ws_data.column_dimensions[get_column_letter(
            col_idx)].width = adjusted_width

    # Freeze header row
    ws_data.freeze_panes = 'A2'

    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    create_summary_sheet(ws_summary, df, duplicate_count)

    return wb


def create_summary_sheet(ws, df, duplicate_count):
    """
    Create summary sheet with comprehensive statistics.
    """
    row = 1

    # === OVERVIEW ===
    ws.cell(row=row, column=1, value='CHARGE SCRUB SUMMARY').font = Font(
        bold=True, size=14)
    row += 2

    # Date range
    if 'SERVICEDATE' in df.columns:
        dates = pd.to_datetime(df['SERVICEDATE'], errors='coerce')
        min_date = dates.min()
        max_date = dates.max()
        if pd.notna(min_date) and pd.notna(max_date):
            ws.cell(row=row, column=1, value='Date Range:')
            ws.cell(row=row, column=2,
                    value=f"{min_date.strftime('%m/%d/%Y')} - {max_date.strftime('%m/%d/%Y')}")
            row += 2

    # === TOTALS ===
    ws.cell(row=row, column=1, value='TOTALS').font = Font(bold=True)
    row += 1

    total = len(df)
    total_patients = df['PATIENTNAME'].nunique(
    ) if 'PATIENTNAME' in df.columns else 'N/A'

    ws.cell(row=row, column=1, value='Total Charges')
    ws.cell(row=row, column=2, value=total)
    row += 1
    ws.cell(row=row, column=1, value='Unique Patients')
    ws.cell(row=row, column=2, value=total_patients)
    row += 1
    ws.cell(row=row, column=1, value='Duplicates')
    ws.cell(row=row, column=2, value=duplicate_count)
    row += 2

    # === RECOMMENDATIONS ===
    ws.cell(row=row, column=1, value='RECOMMENDATION').font = Font(bold=True)
    row += 1

    charge_count = (df['_Recommendation'] == 'CHARGE').sum()
    deny_count = (df['_Recommendation'] == 'DENY').sum()
    review_count = (df['_Recommendation'] == 'REVIEW').sum()

    cell = ws.cell(row=row, column=1, value='CHARGE')
    cell.fill = CHARGE_FILL
    ws.cell(row=row, column=2, value=charge_count)
    row += 1
    cell = ws.cell(row=row, column=1, value='DENY')
    cell.fill = DENY_FILL
    ws.cell(row=row, column=2, value=deny_count)
    row += 1
    cell = ws.cell(row=row, column=1, value='REVIEW')
    cell.fill = REVIEW_FILL
    ws.cell(row=row, column=2, value=review_count)
    row += 2

    # === BY PROVIDER ===
    if 'BILLINGPROVIDER' in df.columns:
        ws.cell(row=row, column=1, value='BY PROVIDER').font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value='Provider').font = Font(bold=True)
        ws.cell(row=row, column=2, value='Charges').font = Font(bold=True)
        ws.cell(row=row, column=3, value='CHARGE').font = Font(bold=True)
        ws.cell(row=row, column=4, value='DENY').font = Font(bold=True)
        ws.cell(row=row, column=5, value='REVIEW').font = Font(bold=True)
        row += 1

        provider_stats = df.groupby('BILLINGPROVIDER').agg(
            Charges=('BILLINGPROVIDER', 'count'),
            CHARGE=('_Recommendation', lambda x: (x == 'CHARGE').sum()),
            DENY=('_Recommendation', lambda x: (x == 'DENY').sum()),
            REVIEW=('_Recommendation', lambda x: (x == 'REVIEW').sum())
        ).reset_index()

        for _, prow in provider_stats.iterrows():
            ws.cell(row=row, column=1, value=prow['BILLINGPROVIDER'])
            ws.cell(row=row, column=2, value=prow['Charges'])
            ws.cell(row=row, column=3, value=prow['CHARGE'])
            ws.cell(row=row, column=4, value=prow['DENY'])
            ws.cell(row=row, column=5, value=prow['REVIEW'])
            row += 1
        row += 1

    # === BY CPT CODE ===
    if 'PROCEDURECODE' in df.columns:
        ws.cell(row=row, column=1, value='BY CPT CODE').font = Font(bold=True)
        row += 1
        ws.cell(row=row, column=1, value='CPT').font = Font(bold=True)
        ws.cell(row=row, column=2, value='Count').font = Font(bold=True)
        ws.cell(row=row, column=3, value='CHARGE').font = Font(bold=True)
        ws.cell(row=row, column=4, value='DENY').font = Font(bold=True)
        ws.cell(row=row, column=5, value='REVIEW').font = Font(bold=True)
        row += 1

        cpt_stats = df.groupby('PROCEDURECODE').agg(
            Count=('PROCEDURECODE', 'count'),
            CHARGE=('_Recommendation', lambda x: (x == 'CHARGE').sum()),
            DENY=('_Recommendation', lambda x: (x == 'DENY').sum()),
            REVIEW=('_Recommendation', lambda x: (x == 'REVIEW').sum())
        ).reset_index().sort_values('Count', ascending=False)

        for _, crow in cpt_stats.iterrows():
            ws.cell(row=row, column=1, value=crow['PROCEDURECODE'])
            ws.cell(row=row, column=2, value=crow['Count'])
            ws.cell(row=row, column=3, value=crow['CHARGE'])
            ws.cell(row=row, column=4, value=crow['DENY'])
            ws.cell(row=row, column=5, value=crow['REVIEW'])
            row += 1

    # Autofit columns
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10


def lambda_handler(event, context):
    """
    AWS Lambda function to convert CSV/TXT files to Excel format.
    Expects a multipart/form-data POST request with a file.
    Returns the converted Excel file as base64-encoded binary.
    """
    try:
        # Parse the incoming request body
        content_type = event.get('headers', {}).get(
            'content-type', '') or event.get('headers', {}).get('Content-Type', '')
        body = event.get('body', '')
        is_base64 = event.get('isBase64Encoded', False)

        # Decode base64 body if needed
        if is_base64:
            body = base64.b64decode(body)
        else:
            body = body.encode('utf-8') if isinstance(body, str) else body

        # Parse multipart form data
        file_content, filename = parse_multipart(body, content_type)

        if not file_content:
            return error_response(400, "No file provided")

        # Detect delimiter
        delimiter = detect_delimiter(file_content)

        # If delimiter is ^, remove trailing ^ from each line to prevent empty column
        if delimiter == '^':
            lines = file_content.split('\n')
            cleaned_lines = [line.rstrip('^') for line in lines]
            file_content = '\n'.join(cleaned_lines)

        # Read CSV into pandas DataFrame
        df = pd.read_csv(io.StringIO(file_content),
                         delimiter=delimiter, dtype=str)

        if df.empty:
            return error_response(400, "File is empty")

        # Remove trailer rows (Epic exports often have a "T" row at the end with record count)
        # These rows have mostly empty values - filter out rows where first column is "T"
        if 'GUARANTORACCOUNT' in df.columns:
            df = df[df['GUARANTORACCOUNT'] != 'T']

        # Apply charge scrubbing (adds _Recommendation, _Notes columns, detects duplicates)
        df, duplicate_count = apply_charge_scrub(df)

        # Create Excel with scrub results
        workbook = create_excel_with_scrub(df, duplicate_count)

        # Save to buffer
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)

        # Generate output filename
        output_filename = filename.rsplit(
            '.', 1)[0] + '.xlsx' if '.' in filename else filename + '.xlsx'

        # Return the Excel file
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': f'attachment; filename="{output_filename}"',
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Headers': 'Content-Type',
                'Access-Control-Allow-Methods': 'POST, OPTIONS',
                'Access-Control-Expose-Headers': 'Content-Disposition'
            },
            'body': base64.b64encode(excel_buffer.read()).decode('utf-8'),
            'isBase64Encoded': True
        }

    except Exception as e:
        print(f"Error: {str(e)}")
        return error_response(500, f"Conversion failed: {str(e)}")


def parse_multipart(body, content_type):
    """
    Parse multipart/form-data to extract file content and filename.
    """
    # Extract boundary from content-type
    boundary = None
    for part in content_type.split(';'):
        part = part.strip()
        if part.startswith('boundary='):
            boundary = part.split('=', 1)[1].strip('"')
            break

    if not boundary:
        # Try parsing as raw file content
        return body.decode('utf-8', errors='ignore'), 'file.csv'

    # Split body by boundary
    boundary_bytes = f'--{boundary}'.encode()
    parts = body.split(boundary_bytes)

    for part in parts:
        if b'filename=' in part:
            # Extract filename
            header_end = part.find(b'\r\n\r\n')
            if header_end == -1:
                continue

            headers = part[:header_end].decode('utf-8', errors='ignore')
            content = part[header_end + 4:]

            # Remove trailing boundary markers
            if content.endswith(b'--\r\n'):
                content = content[:-4]
            elif content.endswith(b'\r\n'):
                content = content[:-2]

            # Extract filename from headers
            filename = 'file.csv'
            for line in headers.split('\r\n'):
                if 'filename=' in line:
                    start = line.find('filename=') + 9
                    end = line.find(
                        '"', start + 1) if line[start] == '"' else line.find(';', start)
                    if end == -1:
                        end = len(line)
                    filename = line[start:end].strip('"')
                    break

            return content.decode('utf-8', errors='ignore'), filename

    return None, None


def detect_delimiter(content):
    """
    Detect the file delimiter: ^, tab, or comma.
    """
    first_line = content.split('\n')[0] if '\n' in content else content

    caret_count = first_line.count('^')
    comma_count = first_line.count(',')
    tab_count = first_line.count('\t')

    # Return the most common delimiter
    counts = {'^': caret_count, '\t': tab_count, ',': comma_count}
    return max(counts, key=counts.get)


def error_response(status_code, message):
    """
    Return a formatted error response.
    """
    return {
        'statusCode': status_code,
        'headers': {
            'Content-Type': 'application/json',
            'Access-Control-Allow-Origin': '*',
            'Access-Control-Allow-Headers': 'Content-Type',
            'Access-Control-Allow-Methods': 'POST, OPTIONS'
        },
        'body': json.dumps({'error': message})
    }
